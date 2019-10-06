from django.shortcuts import render
import openpyxl


def index(request):
    if 'GET' == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES['excel_file']

        wb = openpyxl.load_workbook(excel_file)

        worksheet = wb['Sheet1']
        print(worksheet)

        excel_data = list()
        # データ部から読み込みたいときは、min_row=2を設定する
        # for row in worksheet.iter_rows(min_row=2):
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        return render(request, 'myapp/index.html', {'excel_data': excel_data})
